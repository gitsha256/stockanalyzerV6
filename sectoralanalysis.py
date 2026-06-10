
import os
import sys
import glob
import re
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# RRG ENGINE (v1 Re-implementation)
# ─────────────────────────────────────────────
def calculate_rrg_coordinates(sector_series: pd.Series, bench_series: pd.Series) -> tuple:
    # ALIGNMENT FIX: Normalize indices to Date only (remove time/timezone)
    sector_series.index = pd.to_datetime(sector_series.index).tz_localize(None).normalize()
    bench_series.index = pd.to_datetime(bench_series.index).tz_localize(None).normalize()

    aligned_df = pd.DataFrame({'Sector': sector_series, 'Bench': bench_series}).dropna()
    
    if len(aligned_df) < 80: # Reduced from 100 to 80 to be more flexible for new users
        print(f"   [DEBUG] Insufficient overlap for RRG: {len(aligned_df)} days found (min 80 required). Defaulting to 100.0")
        return pd.Series(100.0, index=sector_series.index), pd.Series(100.0, index=sector_series.index)
    
    rs = (aligned_df['Sector'] / aligned_df['Bench']) * 100
    ema_rs = rs.ewm(span=60, adjust=False).mean()
    std_rs = rs.rolling(window=60, min_periods=30).std()
    rs_ratio = 100 + 10 * ((rs - ema_rs) / std_rs.replace(0, np.nan)).fillna(0)
    
    d_ratio = rs_ratio.diff().fillna(0)
    ema_d_fast = d_ratio.ewm(span=10, adjust=False).mean()
    ema_d_slow = d_ratio.ewm(span=60, adjust=False).mean()
    std_d = d_ratio.rolling(window=60, min_periods=30).std()
    rs_momentum = 100 + 10 * ((ema_d_fast - ema_d_slow) / std_d.replace(0, np.nan)).fillna(0)
    
    return rs_ratio.reindex(sector_series.index).fillna(100.0), rs_momentum.reindex(sector_series.index).fillna(100.0)

def calculate_velocity_and_heading(ratio_s, mom_s):
    if len(ratio_s) < 2: return 0.0, 0.0
    dx = ratio_s.iloc[-1] - ratio_s.iloc[-2]
    dy = mom_s.iloc[-1] - mom_s.iloc[-2]
    vel = np.sqrt(dx**2 + dy**2)
    heading = (np.degrees(np.arctan2(dy, dx)) + 360) % 360
    return vel, heading

def get_sector_rankings(snap: pd.DataFrame, snapshot_path: str) -> pd.DataFrame:
    # Prefer data_all.csv if using snapshot_all.csv, else prefer data.csv
    if snapshot_path.endswith("snapshot_all.csv"):
        history_file = "data_all.csv" if os.path.exists("data_all.csv") else "data.csv"
    else:
        history_file = "data.csv" if os.path.exists("data.csv") else "data_all.csv"

    if not os.path.exists(history_file):
        raise FileNotFoundError(f"Historical data file ({history_file}) not found. Run main.py first.")
    
    print(f"   Loading history from {history_file}...")
    hist = pd.read_csv(history_file)
    hist['datetime'] = pd.to_datetime(hist['datetime'])
    hist['symbols'] = hist['symbols'].str.upper().str.strip()
    print(f"   [INFO] Local history contains {hist['datetime'].nunique()} unique trading days.")
    
    # Create sector mapping and filter history
    sector_map = snap.set_index('symb')['sect'].to_dict()
    hist['sect'] = hist['symbols'].map(sector_map)
    
    # Synthesize Sector Indices (Equal Weighted)
    print("   Synthesizing sector indices...")
    sector_prices = hist.groupby(['datetime', 'sect'])['close'].mean().unstack().ffill()
    sector_prices.index = pd.to_datetime(sector_prices.index).tz_localize(None).normalize()

    # Get Benchmark
    print("   Fetching Nifty 50 benchmark...")
    try:
        # Check historical coverage requirement
        hist_end = hist['datetime'].max().tz_localize(None).normalize()
        
        bench_df = yf.download("^NSEI", period="2y", progress=False)
        if bench_df.empty:
            raise ValueError("yfinance returned no data")
            
        # Handle potential MultiIndex columns in recent yfinance versions
        if isinstance(bench_df.columns, pd.MultiIndex):
            if 'Close' in bench_df.columns.levels[0]:
                bench = bench_df['Close'].iloc[:, 0]
            else:
                bench = bench_df.iloc[:, 0]
        else:
            bench = bench_df['Close']
            
        if isinstance(bench, pd.DataFrame): bench = bench.squeeze()

        bench.index = pd.to_datetime(bench.index).tz_localize(None).normalize()
        
        # Verify if the benchmark reaches the end of our history
        if bench.index.max() < hist_end:
            raise ValueError(f"Benchmark ends at {bench.index.max().date()}, but history reaches {hist_end.date()}")

        print(f"   [SUCCESS] Nifty 50 benchmark fetched ({len(bench)} days of history).")
    except Exception as e:
        print(f"   [WARN] Benchmark coverage gap ({e}) — using internal market proxy.")
        try:
            # Build a proxy Nifty benchmark by computing the equal-weighted average of daily close prices across ALL symbols
            bench = hist.groupby('datetime')['close'].mean()
            bench.index = pd.to_datetime(bench.index).tz_localize(None).normalize()
            if bench.empty:
                raise ValueError("Market proxy calculation returned no data")
        except Exception as fe:
            raise RuntimeError(f"Critical Error: Failed to fetch Nifty benchmark and failed to build proxy from history. {fe}") from fe

    sector_summary = []
    for sector in sector_prices.columns:
        if pd.isna(sector) or sector == 'Unknown': continue
        
        ratio_s, mom_s = calculate_rrg_coordinates(sector_prices[sector], bench)
        
        r = ratio_s.iloc[-1]
        m = mom_s.iloc[-1]
        vel, head = calculate_velocity_and_heading(ratio_s, mom_s)
        
        # ── Aggregating Breadth Metrics from Snapshot ──
        sec_snap = snap[snap['sect'] == sector]
        if sec_snap.empty: continue

        # 1. RSI Breadth (User Suggestion: % of stocks with RSI >= 50)
        rsi_breadth = (sec_snap['rsi'] >= 50).mean() * 100
        
        # 2. EMA Breadth (% stocks above SMA50)
        ema_breadth = (sec_snap['g050'] == True).mean() * 100
        
        # 3. Delivery Conviction (Average delivery %)
        avg_delivery = sec_snap['DlPer'].mean()

        # 4. Heading Direction Score (Rewards NE movement towards Leading quadrant)
        # Formula: cos(heading - 45 degrees) mapped to 0-100
        heading_score = max(0, np.cos(np.radians(head - 45))) * 100

        # ── Weighted Multi-Factor Rotational Score (Institutional Grade) ──
        # Weights: RSI Breadth(25%), EMA Breadth(30%), Heading(25%), Delivery(20%)
        score = (rsi_breadth * 0.25) + (ema_breadth * 0.30) + (heading_score * 0.25) + (avg_delivery * 0.20)

        # Quadrant Determination
        quad = "LEADING" if r >= 100 and m >= 100 else "WEAKENING" if r >= 100 else "LAGGING" if m < 100 else "IMPROVING"
        
        sector_summary.append({
            'Sector Name': sector,
            'RRG Quadrant': quad,
            'RS-Ratio': round(r, 2),
            'RS-Momentum': round(m, 2),
            'Rotational Score': round(score, 2),
            'Velocity': round(vel, 2),
            'Heading': round(head, 1)
        })

    df = pd.DataFrame(sector_summary).sort_values("Rotational Score", ascending=False)
    print(f"   Sectors  : {len(df)} calculated manually")
    return df

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
DARKSOUL_XLSX     = "sector_rotation_multi_report.xlsx"
OUTPUT_FILE       = "swing_scanner_report.xlsx"

# Sector filter — which quadrants to include
GOOD_QUADRANTS    = {"LEADING", "IMPROVING"}

# Stock filters
MIN_RSI           = 45
MAX_RSI           = 72
MIN_ADX           = 20
MIN_RVOL          = 1.0

BULLISH_PATTERNS  = {
    "Inverse Head and Shoulders", "Double Bottom", "Triple Bottom",
    "Ascending Triangle", "Flag Pattern (Bull Flag)", "Pennant Pattern",
    "Falling Wedge", "Cup and Handle", "Rounding Bottom",
    "Long-term Cup and Handle", "Long-term Rounding Bottom",
    "Rectangle Pattern", "Symmetrical Triangle",
}
BEARISH_PATTERNS  = {
    "Head and Shoulders", "Double Top", "Triple Top",
    "Descending Triangle", "Flag Pattern (Bear Flag)",
    "Rising Wedge", "Rounding Top", "Diamond Top",
}

# ─────────────────────────────────────────────
# STEP 1 — Load snapshot.csv
# ─────────────────────────────────────────────
def load_snapshot(file_path: str) -> pd.DataFrame:
    print(f"   Snapshot : {file_path}")
    df = pd.read_csv(file_path).fillna(np.nan)
    df.columns = df.columns.str.strip()
    if 'symb' in df.columns:
        df['symb'] = df['symb'].astype(str).str.upper().str.strip()
    return df

# ─────────────────────────────────────────────
# STEP 3 — Filter + merge
# ─────────────────────────────────────────────
def build_candidates(snap: pd.DataFrame, sectors: pd.DataFrame) -> tuple:

    # Leading / Improving sectors only
    good_sectors = sectors[sectors['RRG Quadrant'].isin(GOOD_QUADRANTS)]['Sector Name'].tolist()
    print(f"\n   ✅ Good sectors ({len(good_sectors)}): {good_sectors}")
    steps = {}

    if not good_sectors:
        print("   ⚠️  No Leading/Improving sectors today — no trades.")
        return pd.DataFrame(), steps

    df = snap.copy()
    date_col = 'date' if 'date' in df.columns else None
    symb_col = 'symb' if 'symb' in df.columns else None

    # ── Sector filter ──
    sect_col = 'sect' if 'sect' in df.columns else None
    if sect_col:
        df = df[df[sect_col].isin(good_sectors)]
        steps["Filter - Sector"] = df[[c for c in [symb_col, sect_col] if c]].copy()
    print(f"   After sector filter    : {len(df)}")

    # ── Stage 2 only ──
    if 'stge' in df.columns:
        df = df[df['stge'].astype(str).str.contains("Stage 2", na=False)]
        steps["Filter - Stage 2"] = df[[c for c in [symb_col, sect_col, 'stge'] if c]].copy()
    print(f"   After Stage 2 filter   : {len(df)}")

    # ── Above SMA200 ──
    if 'g200' in df.columns:
        # Coerce "True"/"False" strings to actual booleans
        g200_mask = df['g200'].map(lambda x: True if str(x).strip().lower() in ("true", "1", "yes") else False)
        df = df[g200_mask]
        steps["Filter - SMA200"] = df[[c for c in [symb_col, sect_col, 'g200'] if c]].copy()
    print(f"   After SMA200 filter    : {len(df)}")

    # ── RSI range ──
    if 'rsi' in df.columns:
        rsi = pd.to_numeric(df['rsi'], errors='coerce')
        df  = df[(rsi >= MIN_RSI) & (rsi <= MAX_RSI)]
        steps["Filter - RSI"] = df[[c for c in [symb_col, sect_col, 'rsi'] if c]].copy()
    print(f"   After RSI filter       : {len(df)}")

    # ── ADX ──
    if 'adx' in df.columns:
        df = df[pd.to_numeric(df['adx'], errors='coerce').fillna(0) >= MIN_ADX]
        steps["Filter - ADX"] = df[[c for c in [symb_col, sect_col, 'adx'] if c]].copy()
    print(f"   After ADX filter       : {len(df)}")

    # ── Relative Volume ──
    if 'rvol' in df.columns:
        df = df[pd.to_numeric(df['rvol'], errors='coerce').fillna(0) >= MIN_RVOL]
        steps["Filter - RVOL"] = df[[c for c in [symb_col, sect_col, 'rvol'] if c]].copy()
    print(f"   After RVOL filter      : {len(df)}")

    # ── Pattern filter ──
    pat_col = next((c for c in ['mpat', 'patt'] if c in df.columns), None)
    if pat_col:
        df = df[~df[pat_col].isin(BEARISH_PATTERNS)]
        steps["Filter - Pattern"] = df[[c for c in [symb_col, sect_col, pat_col] if c]].copy()
        df = df.copy()
        df['pat_bullish'] = df[pat_col].isin(BULLISH_PATTERNS).astype(int)
    else:
        df['pat_bullish'] = 0
    print(f"   After pattern filter   : {len(df)}")

    if df.empty:
        return df, steps

    # ── Merge sector RRG data ──
    sector_info = sectors[['Sector Name', 'RRG Quadrant', 'RS-Ratio',
                            'RS-Momentum', 'Rotational Score', 'Velocity', 'Heading']].copy()
    sector_info = sector_info.rename(columns={'Sector Name': sect_col})
    df = df.merge(sector_info, on=sect_col, how='left')

    # ── Sort: pat_bullish → rotational score → rvol ──
    sort_cols = [c for c in ['pat_bullish', 'Rotational Score', 'rvol'] if c in df.columns]
    df = df.sort_values(sort_cols, ascending=False)

    # ── Final output columns ──
    keep = []
    for col in ['date', 'symb', sect_col, 'RRG Quadrant', 'Rotational Score', 'RS-Ratio', 'RS-Momentum', 'MT_Zone',
                'clos', 'rsi', 'adx', 'stge', 'rvol', 'DlPer',
                'g200', 'g050', 'zone', 'obv',
                'mpat', 'pcon', 'patt', 'pat_bullish',
                'bbsq', 'bbbw', 'mcdl', 'vola', 'delt',
                'shgh', 'slw', 'tren', 'tstr']:
        if col and col in df.columns:
            keep.append(col)
    return df[keep], steps

# ─────────────────────────────────────────────
# STEP 4 — Styling Layer
# ─────────────────────────────────────────────
def apply_corporate_styling(file_path):
    wb = openpyxl.load_workbook(file_path)
    cell_font = Font(name="Segoe UI", size=10)
    
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        
        # ── Convert to Table ──
        if ws.max_row > 1:
            last_col = get_column_letter(ws.max_column)
            tab_range = f"A1:{last_col}{ws.max_row}"
            # Display name must be unique and contain no spaces
            safe_name = "".join(filter(str.isalnum, sheet_name)) + f"_{i}"
            tab = Table(displayName=f"Table_{safe_name}", ref=tab_range)
            style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Left align headers for better readability when filter icons are active
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='left')

        # Column width auto-fit
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
                cell.font = cell_font
            ws.column_dimensions[column].width = min(max_length + 3, 40)
            
    wb.save(file_path)

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("\n" + "="*60)
    print("🚀 Swing Scanner Bridge v3.0 (Excel Output)")
    print("="*60)

    # 1. Ask for snapshot type
    print("Choose snapshot source:")
    print("1. snapshot.csv")
    print("2. snapshot_all.csv")
    choice = input("Enter choice [default 1]: ").strip()

    if choice == '2':
        suffix_flag = "_all"
        pattern = "*snapshot_all.csv"
    else:
        suffix_flag = ""
        pattern = "*snapshot.csv"

    # 2. Ask for data target
    print("\nOptions: press Enter to latest data, enter a specific date (DD-MM-YYYY),")
    print("a range (DD-MM-YYYY to DD-MM-YYYY), or a custom filename.")
    user_input = input("Target: ").strip()

    targets = []

    if not user_input:
        all_files = glob.glob(pattern)
        if choice != '2':
            all_files = [f for f in all_files if not f.endswith("_all.csv")]
        
        # Filter for dated files only (ignores ma_ or sma_ prefixes)
        files = [f for f in all_files if re.match(r"^\d{2}-\d{2}-\d{2,4}", os.path.basename(f))]
        
        if not files:
            print(f"No files matching {pattern} found in current directory.")
            sys.exit(1)

        # Sort by the date embedded in the filename (DD-MM-YY)
        def get_file_date(fname):
            name = os.path.basename(fname)
            try:
                return datetime.strptime(name[:8], "%d-%m-%y")
            except ValueError:
                try:
                    return datetime.strptime(name[:10], "%d-%m-%Y")
                except ValueError:
                    return datetime.fromtimestamp(0)

        files.sort(key=get_file_date)
        targets.append(files[-1])

    elif "to" in user_input:
        # Handle range
        try:
            start_str, end_str = [d.strip() for d in user_input.split('to')]
            start_dt = datetime.strptime(start_str, '%d-%m-%Y')
            end_dt = datetime.strptime(end_str, '%d-%m-%Y')
            
            curr = start_dt
            while curr <= end_dt:
                fname = f"{curr.strftime('%d-%m-%y')}snapshot{suffix_flag}.csv"
                if os.path.exists(fname):
                    targets.append(fname)
                curr += timedelta(days=1)
        except Exception as e:
            print(f"Error parsing date range: {e}")
            sys.exit(1)

    elif os.path.exists(user_input):
        targets.append(user_input)

    else:
        # Try as single date
        try:
            dt = datetime.strptime(user_input, '%d-%m-%Y')
            fname = f"{dt.strftime('%d-%m-%y')}snapshot{suffix_flag}.csv"
            if os.path.exists(fname):
                targets.append(fname)
            else:
                print(f"File {fname} not found.")
                sys.exit(1)
        except ValueError:
            print(f"Invalid input format or file not found: {user_input}")
            sys.exit(1)

    if not targets:
        print("No valid snapshot files identified to process.")
        sys.exit(1)

    for snapshot_path in targets:
        print(f"\n📂 Loading files from {snapshot_path}...")
        try:
            snap    = load_snapshot(snapshot_path)
            sectors = get_sector_rankings(snap, snapshot_path)
        except Exception as e:
            print(f"❌ Error during manual RRG calculation for {snapshot_path}: {e}")
            continue

        print("\n📊 Sector RRG snapshot (darksoul):")
        print(sectors[['Sector Name', 'RRG Quadrant', 'RS-Ratio',
                       'RS-Momentum', 'Rotational Score']].to_string(index=False))
        print("\n🔍 Filtering swing candidates...")
        candidates, filter_steps = build_candidates(snap, sectors)

        # Determine output filename based on snapshot filename to avoid overwrites
        base_name_file = os.path.splitext(os.path.basename(snapshot_path))[0]
        final_output_file = f"sector_report_{base_name_file}.xlsx"

        # Write to Excel
        print(f"\n💾 Compiling Master Report: {final_output_file}")
        with pd.ExcelWriter(final_output_file, engine='openpyxl') as writer:
            candidates.to_excel(writer, sheet_name="Swing Candidates", index=False)
            sectors.to_excel(writer, sheet_name="Sector RRG Snapshot", index=False)
            # Add bifurcated filter sheets
            for sheet_name, step_df in filter_steps.items():
                step_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        if candidates.empty:
            print("\n⚠️  No final candidates found, but filter progression saved to Excel.")
        else:
            print(f"✅ {len(candidates)} candidates found.")

        apply_corporate_styling(final_output_file)

        print(f"✅ Report saved and styled: {final_output_file}")
        print(f"\n🏆 Top 10 Candidates ({base_name_file}):")
        sect_col = 'sect'
        print(candidates[['symb', sect_col,
                           'RRG Quadrant', 'Rotational Score',
                           'clos', 'rsi', 'adx', 'mpat']].head(10).to_string(index=False))
        print("\n" + "="*60)
     
if __name__ == "__main__":
    main()
