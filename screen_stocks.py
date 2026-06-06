import pandas as pd
import numpy as np
import sys
import os
import re
import glob
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# CONFIG — tweak weights/thresholds here without touching logic
# ─────────────────────────────────────────────────────────────

INTRADAY_CFG = {
    # ── Hard filters (ALL must pass) ──
    "stage":            "Stage 2",          # stge must contain this
    "rsi_min":          45.0,
    "rsi_max":          72.0,
    "adx_min":          18.0,
    "rvol_min":         0.8,                # relative volume floor
    "ascr_min":         20.0,               # ₹ liquidity floor
    "trend_allow":      ["Uptrend", "Sideways"],
    "above_sma200":     True,

    # ── Base scoring weights ──
    "w_rvol":           25.0,   # volume surge — #1 intraday signal
    "w_adx":            1.5,    # trend strength
    "w_rsi":            0.5,    # momentum (mild — avoids chasing overbought)
    "w_ascr_log":       5.0,    # log(ascr) liquidity — log-scaled prevents mega-caps dominating
    "w_positive_day":   10.0,   # bonus: chan > 0
    "w_uptrend":        8.0,    # bonus: tren == Uptrend
    "w_strong_trend":   8.0,    # bonus: tstr == Strong
    "w_bbbreakout":     5.0,    # bonus: bbup == True

    # ── Test Feature Weights ──
    "w_bbbw_low":       15.0,   # rewards tighter bandwidth (volatility contraction)
    "w_mtf_align":      6.0,    # rsi > 55 AND wrsi > 60
    "w_weak_penalty":   -10.0,  # penalty for tstr == Weak

    # ── New indicator weights (all pre-computed in snapshot) ──
    "w_supertrend":     10.0,   # bonus: SUPERTd == +1  (uptrend confirmed)
    "w_cmf":            8.0,    # CMF × weight  (negative CMF penalises score)
    "w_squeeze_on":     6.0,    # bonus: SQZ_ON == 1  (coiling before breakout)
    "w_stoch_setup":    5.0,    # bonus: STOCHk < 80 AND %k > %d  (bullish, not OB)
    "w_efi_positive":   3.0,    # bonus: EFI_13 > 0  (buying force present)

    "top_n":            10,
}

SWING_CFG = {
    # ── Hard filters ──
    "stage":            "Stage 2",
    "wrsi_min":         50.0,
    "wrsi_max":         75.0,               # cap: >75 = weekly exhaustion risk
    "adx_min":          18.0,
    "dlper_min":        35.0,               # delivery % — filters speculative churn
    "delt_max":         25.0,               # within 25% of 52W high
    "above_sma200":     True,
    "above_sma050":     True,
    "trend_require":    "Uptrend",

    # ── Base scoring weights ──
    "w_wrsi":           1.0,    # weekly RSI momentum
    "w_adx":            1.0,    # trend strength
    "w_delt_prox":      0.5,    # proximity to 52W high: (25−delt) → higher = closer
    "w_pcon":           0.3,    # chart pattern confidence
    "w_strong_trend":   15.0,   # largest bonus — swing lives/dies on trend quality
    "w_bbsqueeze":      10.0,   # BB squeeze: compressed vol = breakout potential
    "w_premium_zone":   8.0,    # bonus: zone Premium or Near Premium
    "w_ascr_log":       2.0,    # liquidity (log-scaled)

    # ── Test Feature Weights ──
    "w_bbbw_low":       20.0,   # reward low bbbw (compressed volatility)
    "w_mtf_align":      8.0,    # dual timeframe momentum bonus
    "w_weak_penalty":   -12.0,  # penalty for weak trend
    "w_weekly_stage_align": 10.0, # adds bonus for MTF Weinstein confluence

    # ── New indicator weights ──
    "w_supertrend":     12.0,   # bonus: SUPERTd == +1  (multi-day confirmation)
    "w_cmf":            10.0,   # CMF × weight  (institutional accumulation)
    "w_squeeze_on":     8.0,    # bonus: SQZ_ON == 1  (imminent swing move)
    "w_willr_pullback": 6.0,    # bonus: WILLR < -60  (pulled back, entry timing)
    "w_ema21_support":  5.0,    # bonus: clos > EMA_21  (dynamic support intact)
    "w_stoch_setup":    4.0,    # bonus: STOCHk < 70 AND %k > %d
    "w_rsi2_pullback":  7.0,    # bonus for RSI_2 < 25

    "top_n":            10,
}

# ── Output column lists (missing cols auto-skipped) ──
OUTPUT_COLS_INTRADAY = [
    "symb", "clos", "chan", "rvol", "arnk", "ascr",
    "rsi", "adx", "tren", "tstr", "vola",
    "zone", "MT_Zone", "delt", "bbup", "bbsq",
    "SUPERTd_7_3.0", "CMF_20", "SQZ_ON", "STOCHk_14_3_3", "EFI_13",
    "mpat", "pcon", "sect", "score",
]

OUTPUT_COLS_SWING = [
    "symb", "clos", "chan", "wrsi", "ws30", "DlPer",
    "delt", "adx", "tstr", "vola", "zone", "MT_Zone",
    "SUPERTd_7_3.0", "CMF_20", "SQZ_ON", "WILLR_14", "EMA_21",
    "STOCHk_14_3_3",
    "mpat", "pcon", "xpat", "sect", "score",
]


# ─────────────────────────────────────────────
# LOAD & VALIDATE
# ─────────────────────────────────────────────
def load_csv(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        sys.exit(f"[ERROR] File not found: {path}")
    df = pd.read_csv(path)

    # core columns that must exist
    required = {
        "stge", "g200", "g050", "rsi", "wrsi", "adx", "rvol",
        "ascr", "chan", "tren", "tstr", "vola", "zone", "delt",
        "DlPer", "bbup", "bbsq", "bbbw", "mpat", "pcon", "ws30", "clos", "sect", "MT_Zone"
    }
    missing = required - set(df.columns)
    if missing:
        sys.exit(f"[ERROR] Missing required columns: {missing}")

    # new indicator columns — warn if absent but don't abort
    new_cols = [
        "CMF_20", "SUPERTd_7_3.0", "STOCHk_14_3_3", "STOCHd_14_3_3",
        "EMA_21", "SQZ_ON", "SQZ_OFF", "WILLR_14", "EFI_13", "RSI_2",
    ]
    missing_new = [c for c in new_cols if c not in df.columns]
    if missing_new:
        print(f"[WARN] New indicator columns not found (scoring will skip them): {missing_new}")

    # ── type coercions ──
    bool_cols = ["g200", "g050", "g020", "bbup", "bbsq"]
    for col in bool_cols:
        if col in df.columns:
            df[col] = df[col].map(
                lambda x: True if str(x).strip().lower() in ("true", "1", "yes") else False
            )

    int_cols = ["SQZ_ON", "SQZ_OFF", "SQZ_NO"]
    for col in int_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    float_cols = [
        "CMF_20", "SUPERTd_7_3.0", "STOCHk_14_3_3", "STOCHd_14_3_3",
        "EMA_21", "WILLR_14", "EFI_13", "RSI_2", "bbbw",
    ]
    for col in float_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # fill optional / nullable cols
    df["pcon"]  = pd.to_numeric(df["pcon"], errors="coerce").fillna(0)
    df["xpat"]  = df.get("xpat", pd.Series([""] * len(df))).fillna("")
    df["mpat"]  = df["mpat"].fillna("No Pattern")

    return df


def parse_date_from_filename(path: str) -> str:
    """Extract display date from filename like 29-05-26snapshot.csv → 29 May 2026"""
    name = os.path.basename(path)
    m = re.match(r"(\d{2})-(\d{2})-(\d{2})snapshot", name)
    if m:
        dd, mm, yy = m.groups()
        try:
            dt = datetime.strptime(f"{dd}-{mm}-20{yy}", "%d-%m-%Y")
            return dt.strftime("%d %b %Y")
        except ValueError:
            pass
    return name


# ─────────────────────────────────────────────
# INTRADAY FILTER + SCORE
# ─────────────────────────────────────────────
def screen_intraday(df: pd.DataFrame, cfg: dict) -> tuple[pd.DataFrame, int]:
    """
    Hard filters:
      Stage 2 | g200 | RSI 45-72 | ADX≥18 | RVol≥0.8 | ascr≥20 | Trend∈{Up,Sideways}

    Base score:
      RVol×25  ADX×1.5  RSI×0.5  logAscr×5
      +10 positive day  +8 Uptrend  +8 Strong  +5 BB breakout

    New indicator score:
      +10  SUPERTd == +1          (trend direction confirmed)
      CMF×8                       (money flow; negative penalises)
      +6   SQZ_ON                 (coiling = breakout imminent)
      +5   STOCHk<80 AND k>d      (bullish setup, not overbought)
      +3   EFI_13 > 0             (buying force present)
    """
    c = cfg
    mask = (
        df["stge"].str.contains(c["stage"], na=False) &
        (df["g200"] == True) &
        df["rsi"].between(c["rsi_min"], c["rsi_max"]) &
        (df["adx"]  >= c["adx_min"]) &
        (df["rvol"] >= c["rvol_min"]) &
        (df["ascr"] >= c["ascr_min"]) &
        df["tren"].isin(c["trend_allow"])
    )
    pool = df[mask].copy()
    pool_size = len(pool)

    # ── base score ──
    pool["score"] = (
        pool["rvol"]                             * c["w_rvol"] +
        pool["adx"]                              * c["w_adx"] +
        pool["rsi"]                              * c["w_rsi"] +
        np.log1p(pool["ascr"])                   * c["w_ascr_log"] +
        (pool["chan"] > 0).astype(int)           * c["w_positive_day"] +
        (pool["tren"] == "Uptrend").astype(int)  * c["w_uptrend"] +
        (pool["tstr"] == "Strong").astype(int)   * c["w_strong_trend"] +
        (pool["bbup"] == True).astype(int)       * c["w_bbbreakout"]
    )

    # ── EMA21 Distance Scoring ──
    if "EMA_21" in pool.columns:
        ema_dist = ((pool["clos"] - pool["EMA_21"]) / pool["EMA_21"]) * 100
        pool["score"] += np.select(
            [
                (ema_dist >= 0) & (ema_dist <= 5),
                (ema_dist > 5) & (ema_dist <= 10),
                (ema_dist > 10) & (ema_dist <= 15),
                (ema_dist > 15)
            ],
            [10.0, 5.0, 0.0, -10.0],
            default=0
        )

    # ── Multi-Timeframe Alignment ──
    mtf_mask = (pool["rsi"] > 55) & (pool.get("wrsi", 0) > 60)
    pool["score"] += mtf_mask.astype(int) * c["w_mtf_align"]

    # ── Low BBBW Reward (Volatility Contraction) ──
    if "bbbw" in pool.columns:
        pool["score"] += (0.2 - pool["bbbw"]).clip(lower=0) * c["w_bbbw_low"]

    # ── Weak Trend Penalty ──
    pool["score"] += (pool["tstr"] == "Weak").astype(int) * c["w_weak_penalty"]

    # ── Weekly Stage Alignment ──
    if "stge_w" in pool.columns:
        pool["score"] += (pool["stge_w"] == "Stage 2 (Uptrend)").astype(int) * c.get("w_weekly_stage_align", 0)

    # ── new indicator scores ──
    if "SUPERTd_7_3.0" in pool.columns:
        pool["score"] += (pool["SUPERTd_7_3.0"] == 1).astype(int) * c["w_supertrend"]

    if "CMF_20" in pool.columns:
        # clip to ±1 just in case; negative CMF subtracts from score
        pool["score"] += pool["CMF_20"].clip(-1, 1) * c["w_cmf"]

    if "SQZ_ON" in pool.columns:
        pool["score"] += pool["SQZ_ON"] * c["w_squeeze_on"]

    if "STOCHk_14_3_3" in pool.columns and "STOCHd_14_3_3" in pool.columns:
        stoch_ok = (
            (pool["STOCHk_14_3_3"] < 80) &
            (pool["STOCHk_14_3_3"] > pool["STOCHd_14_3_3"])
        )
        pool["score"] += stoch_ok.astype(int) * c["w_stoch_setup"]

    if "EFI_13" in pool.columns:
        pool["score"] += (pool["EFI_13"] > 0).astype(int) * c["w_efi_positive"]

    pool["score"] = pool["score"].round(2)
    cols = [col for col in OUTPUT_COLS_INTRADAY if col in pool.columns]
    return pool.sort_values("score", ascending=False).head(c["top_n"])[cols], pool_size


# ─────────────────────────────────────────────
# SWING FILTER + SCORE
# ─────────────────────────────────────────────
def screen_swing(df: pd.DataFrame, cfg: dict) -> tuple[pd.DataFrame, int]:
    """
    Hard filters:
      Stage 2 | g200 | g050 | wRSI 50-75 | ADX≥18
      DlPer≥35 | delt≤25 | tren==Uptrend

    Base score:
      wRSI×1  ADX×1  DlPer×0.3  proximity×0.5  pcon×0.3
      +15 Strong  +10 BBsqueeze  +8 Premium zone  logAscr×2

    New indicator score:
      +12  SUPERTd == +1          (multi-day trend confirmation)
      CMF×10                      (institutional accumulation; negative penalises)
      +8   SQZ_ON                 (compressed vol = imminent swing)
      +6   WILLR < -60            (pulled back to entry zone)
      +5   clos > EMA_21          (dynamic support intact)
      +4   STOCHk<70 AND k>d      (bullish momentum setup)
    """
    c = cfg
    mask = (
        df["stge"].str.contains(c["stage"], na=False) &
        (df["g200"] == True) &
        (df["g050"] == True) &
        df["wrsi"].between(c["wrsi_min"], c["wrsi_max"]) &
        (df["adx"]   >= c["adx_min"]) &
        (df["DlPer"] >= c["dlper_min"]) &
        (df["delt"]  <= c["delt_max"]) &
        (df["tren"]  == c["trend_require"])
    )
    pool = df[mask].copy()
    pool_size = len(pool)

    # ── base score ──
    pool["score"] = (
        pool["wrsi"]                                               * c["w_wrsi"] +
        pool["adx"]                                                * c["w_adx"] +
        (c["delt_max"] - pool["delt"]).clip(lower=0)               * c["w_delt_prox"] +
        pool["pcon"]                                               * c["w_pcon"] +
        (pool["tstr"] == "Strong").astype(int)                     * c["w_strong_trend"] +
        (pool["bbsq"] == True).astype(int)                         * c["w_bbsqueeze"] +
        pool["zone"].isin(["Premium", "Near Premium"]).astype(int) * c["w_premium_zone"] +
        np.log1p(pool["ascr"])                                     * c["w_ascr_log"]
    )

    # ── Tiered Delivery % Scoring ──
    pool["score"] += np.select(
        [
            pool["DlPer"] >= 80,
            pool["DlPer"] >= 65,
            pool["DlPer"] >= 50,
            pool["DlPer"] >= 35
        ],
        [12.0, 8.0, 5.0, 2.0],
        default=0
    )

    # ── EMA21 Distance Scoring ──
    if "EMA_21" in pool.columns:
        ema_dist = ((pool["clos"] - pool["EMA_21"]) / pool["EMA_21"]) * 100
        pool["score"] += np.select(
            [
                (ema_dist >= 0) & (ema_dist <= 5),
                (ema_dist > 5) & (ema_dist <= 10),
                (ema_dist > 10) & (ema_dist <= 15),
                (ema_dist > 15)
            ],
            [10.0, 5.0, 0.0, -10.0],
            default=0
        )

    # ── Low BBBW Reward ──
    if "bbbw" in pool.columns:
        pool["score"] += (0.15 - pool["bbbw"]).clip(lower=0) * c["w_bbbw_low"]

    # ── Multi-Timeframe Alignment ──
    mtf_mask = (pool["rsi"] > 55) & (pool["wrsi"] > 60)
    pool["score"] += mtf_mask.astype(int) * c["w_mtf_align"]

    # ── Weak Trend Penalty ──
    pool["score"] += (pool["tstr"] == "Weak").astype(int) * c["w_weak_penalty"]

    # ── new indicator scores ──
    if "SUPERTd_7_3.0" in pool.columns:
        pool["score"] += (pool["SUPERTd_7_3.0"] == 1).astype(int) * c["w_supertrend"]

    if "CMF_20" in pool.columns:
        pool["score"] += pool["CMF_20"].clip(-1, 1) * c["w_cmf"]

    if "SQZ_ON" in pool.columns:
        pool["score"] += pool["SQZ_ON"] * c["w_squeeze_on"]

    if "WILLR_14" in pool.columns:
        pool["score"] += (pool["WILLR_14"] < -60).astype(int) * c["w_willr_pullback"]

    if "EMA_21" in pool.columns:
        pool["score"] += (pool["clos"] > pool["EMA_21"]).astype(int) * c["w_ema21_support"]

    # ── RSI_2 Pullback Bonus ──
    if "RSI_2" in pool.columns:
        pullback_mask = (pool["RSI_2"] < 25) & (pool["stge"].str.contains("Stage 2", na=False)) & \
                        (pool["tstr"].isin(["Strong", "Moderate"]))
        pool["score"] += pullback_mask.astype(int) * c["w_rsi2_pullback"]

    if "STOCHk_14_3_3" in pool.columns and "STOCHd_14_3_3" in pool.columns:
        stoch_ok = (
            (pool["STOCHk_14_3_3"] < 70) &
            (pool["STOCHk_14_3_3"] > pool["STOCHd_14_3_3"])
        )
        pool["score"] += stoch_ok.astype(int) * c["w_stoch_setup"]

    pool["score"] = pool["score"].round(2)
    cols = [col for col in OUTPUT_COLS_SWING if col in pool.columns]
    return pool.sort_values("score", ascending=False).head(cfg["top_n"])[cols], pool_size


# ─────────────────────────────────────────────
# STYLING ENGINE
# ─────────────────────────────────────────────
def apply_professional_formatting(file_path, df):
    """Exports dataframe to Excel with TableStyleLight9 and professional color scales."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Screened Picks"

    # Load data into worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    if ws.max_row <= 1:
        wb.save(file_path)
        return

    # 1. Convert to Excel Table
    last_col = get_column_letter(ws.max_column)
    tab = Table(displayName="PicksTable", ref=f"A1:{last_col}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # 2. Setup References
    headers = [cell.value for cell in ws[1]]
    def get_col_ref(name):
        try:
            idx = headers.index(name) + 1
            letter = get_column_letter(idx)
            return f"{letter}2:{letter}{ws.max_row}"
        except ValueError: return None

    def get_col_let(name):
        try:
            return get_column_letter(headers.index(name) + 1)
        except ValueError: return None

    # 3. Conditional Formatting Rules
    # Red-Yellow-Green Scales (High is Good)
    for col in ["chan", "rvol", "ascr", "adx", "score", "DlPer", "STOCHk_14_3_3", "EFI_13", "pcon", "CMF_20", "volu", "obv"]:
        ref = get_col_ref(col)
        if ref:
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            ))

    # Inverse Scales (High is Bad/Risk)
    for col in ["vola", "delt", "arnk"]:
        ref = get_col_ref(col)
        if ref:
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type='min', start_color='63BE7B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='F8696B'
            ))

    # Signal Highlighting (Booleans & Numeric Flags)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    blue_fill  = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid')
    orange_fill = PatternFill(start_color='FCAA67', end_color='FCAA67', fill_type='solid')
    f_dark_green = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    f_blue = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    f_dark_red = PatternFill(start_color='9C0006', end_color='9C0006', fill_type='solid')

    for col in ["bbup", "bbdn", "bbsq", "vspk", "g200", "g050", "g020"]:
        ref = get_col_ref(col)
        if ref:
            # Match literal text "True" or "False" as requested
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"True"'], fill=green_fill))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"TRUE"'], fill=green_fill))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['TRUE'], fill=green_fill))
            
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"False"'], fill=red_fill))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"FALSE"'], fill=red_fill))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['FALSE'], fill=red_fill))

    # Numeric flags remain on their own logic
    for col in ["SUPERTd_7_3.0", "SQZ_ON"]:
        ref = get_col_ref(col)
        if ref:
            ws.conditional_formatting.add(ref, CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            ws.conditional_formatting.add(ref, CellIsRule(operator='lessThanOrEqual', formula=['0'], fill=red_fill))

    # Categorical logic
    for col, mapping in [
        ("tren", {"Uptrend": green_fill, "Downtrend": red_fill, "Sideways": yellow_fill}),
        ("tstr", {"Strong": green_fill, "Weak": red_fill, "Moderate": yellow_fill}),
        ("zone", {
            "Premium": red_fill,
            "Discount": green_fill, 
            "Equilibrium": yellow_fill,
            "Near Discount": f_dark_green,
            "Near Premium": f_dark_red
        }),
        ("stge", {
            "Stage 2 (Uptrend)": green_fill,
            "Stage 4 (Downtrend)": red_fill,
            "Stage 1/3 (Neutral)": yellow_fill
        }),
        ("MT_Zone", {
            "Premium": red_fill,
            "Discount": green_fill,
            "Equilibrium": yellow_fill,
            "Near Discount": f_dark_green,
            "Near Premium": f_dark_red
        }),
        ("screener_type", {"Intraday": blue_fill, "Swing": PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')})
    ]:
        ref = get_col_ref(col)
        if ref:
            for val, fill in mapping.items():
                # Apply white font for dark backgrounds
                font = Font(color='FFFFFF') if fill in [f_dark_green, f_dark_red] else None
                ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=[f'"{val}"'], fill=fill, font=font))

    # 4. Momentum Specific Mappings (wrsi and rsi/rsi2)
    f_blue = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    f_dark_green = PatternFill(start_color='375623', end_color='375623', fill_type='solid')

    # wrsi: >80 Blue, 70-80 Dark Green, 60-70 Green, 50-60 Yellow, 40-50 Orange, <40 Red
    ref_wrsi = get_col_ref("wrsi")
    if ref_wrsi:
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='greaterThan', formula=['80'], fill=f_blue, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['70', '80'], fill=f_dark_green, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['60', '70'], fill=green_fill))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['50', '60'], fill=yellow_fill))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['40', '50'], fill=orange_fill))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='lessThan', formula=['40'], fill=red_fill))

    # rsi / RSI_2: >80 Red, 70-80 Orange, 40-70 Yellow, 20-40 Green, <20 Dark Green
    for col in ["rsi", "RSI_2"]:
        ref_rsi = get_col_ref(col)
        if ref_rsi:
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='greaterThan', formula=['80'], fill=red_fill))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='between', formula=['70', '80'], fill=orange_fill))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='between', formula=['40', '70'], fill=yellow_fill))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='between', formula=['20', '40'], fill=green_fill))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='lessThan', formula=['20'], fill=f_dark_green, font=Font(color='FFFFFF')))

    # 5. Williams %R Specific Mapping (6-Tier Momentum Scale)
    ref = get_col_ref("WILLR_14")
    if ref:
        f_extreme = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid') # Blue
        f_vstrong = PatternFill(start_color='375623', end_color='375623', fill_type='solid') # Dark Green
        f_bullish = PatternFill(start_color='63BE7B', end_color='63BE7B', fill_type='solid') # Green
        f_neutral = PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid') # Yellow
        f_weak    = PatternFill(start_color='FCAA67', end_color='FCAA67', fill_type='solid') # Orange
        f_red     = PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid') # Red

        ws.conditional_formatting.add(ref, CellIsRule(operator='greaterThan', formula=['-10'], fill=f_extreme, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-20', '-10'], fill=f_vstrong, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-50', '-20'], fill=f_bullish))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-80', '-50'], fill=f_neutral))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-90', '-80'], fill=f_weak))
        ws.conditional_formatting.add(ref, CellIsRule(operator='lessThan', formula=['-90'], fill=f_red))

    # 6. Trend Proximity and Stochastic Cross Logic
    c_let = get_col_let("clos")
    
    # SUPERT_7_3.0 Trend Strength Mapping
    s_let, s_ref = get_col_let("SUPERT_7_3.0"), get_col_ref("SUPERT_7_3.0")
    if c_let and s_let and s_ref:
        dist = f"((${c_let}2-${s_let}2)/${s_let}2)*100" # relative row ref (2 without $)
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"{dist}>5"], fill=f_blue, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>2,{dist}<=5)"], fill=f_dark_green, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>0,{dist}<=2)"], fill=green_fill))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>=-2,{dist}<=0)"], fill=yellow_fill))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>=-5,{dist}<-2)"], fill=orange_fill))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"{dist}<-5"], fill=red_fill))

    # EMA_21 Trend Strength Mapping
    e_let, e_ref = get_col_let("EMA_21"), get_col_ref("EMA_21")
    if c_let and e_let and e_ref:
        dist = f"((${c_let}2-${e_let}2)/${e_let}2)*100"
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"{dist}>8"], fill=f_blue, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>3,{dist}<=8)"], fill=f_dark_green, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>0,{dist}<=3)"], fill=green_fill))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>-3,{dist}<=0)"], fill=yellow_fill))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>=-8,{dist}<-3)"], fill=orange_fill))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"{dist}<-8"], fill=red_fill))

    # ws30 Trend Strength Mapping
    w_let, w_ref = get_col_let("ws30"), get_col_ref("ws30")
    if c_let and w_let and w_ref:
        dist = f"((${c_let}2-${w_let}2)/${w_let}2)*100"
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"{dist}>15"], fill=f_blue, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>5,{dist}<=15)"], fill=f_dark_green, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>0,{dist}<=5)"], fill=green_fill))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>=-5,{dist}<=0)"], fill=yellow_fill))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>=-15,{dist}<-5)"], fill=orange_fill))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"{dist}<-15"], fill=red_fill))

    # Stochastic K vs D Crossing and Overbought/Oversold Logic
    k_let, k_ref = get_col_let("STOCHk_14_3_3"), get_col_ref("STOCHk_14_3_3")
    d_let, d_ref = get_col_let("STOCHd_14_3_3"), get_col_ref("STOCHd_14_3_3")
    if k_let and d_let and k_ref:
        # Apply to K and D columns if both exist
        for ref in [k_ref, d_ref] if d_ref else [k_ref]:
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"AND(${k_let}2>${d_let}2,${k_let}2>80)"], fill=f_blue, font=Font(color='FFFFFF')))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"AND(${k_let}2<${d_let}2,${k_let}2<20)"], fill=red_fill))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"${k_let}2>${d_let}2"], fill=green_fill))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"${k_let}2<${d_let}2"], fill=orange_fill))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"ABS(${k_let}2-${d_let}2)<=2"], fill=yellow_fill))

    # 7. Global Refinements
    for i, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(i)
        # Auto-fit Column Width
        max_length = max([len(str(cell.value)) for cell in col])
        ws.column_dimensions[col_letter].width = min(max_length + 2, 35)
        # Bold Headers
        ws.cell(row=1, column=i).font = Font(bold=True)
        # Left align headers to prevent overlap with filter icons
        ws.cell(row=1, column=i).alignment = Alignment(horizontal='left')

    wb.save(file_path)


# ─────────────────────────────────────────────
# DISPLAY HELPERS
# ─────────────────────────────────────────────
def print_section(title: str, df: pd.DataFrame, pool_size: int):
    sep = "═" * 110
    print(f"\n{sep}")
    print(f"  {title}  (pool after filters: {pool_size} stocks, showing top {len(df)})")
    print(sep)
    if df.empty:
        print("  No candidates matched the criteria. Try relaxing the filters in CFG.")
        return
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 220)
    pd.set_option("display.float_format", "{:.2f}".format)
    print(df.to_string(index=False))


def print_logic_summary():
    print("""
╔══════════════════════════════════════════════════════════════════════════╗
║              SCREENING LOGIC — SUMMARY                                  ║
╠══════════════════════════════════════════════════════════════════════════╣
║  INTRADAY                                                                ║
║  Hard : Stage2 | g200 | RSI 45-72 | ADX≥18 | RVol≥0.8                  ║
║         ascr≥20 | Trend∈{Uptrend,Sideways}                              ║
║  Base : RVol(×25) ADX(×1.5) RSI(×0.5) logAscr(×5)                      ║
║         +10 posDay  +8 Uptrend  +8 Strong  +5 BBbreakout                ║
║  New  : +10 SUPERTd=+1  CMF×8  +6 SQZ_ON  +5 Stoch setup  +3 EFI>0    ║
╠══════════════════════════════════════════════════════════════════════════╣
║  SWING (weekly)                                                          ║
║  Hard : Stage2 | g200 | g050 | wRSI 50-75 | ADX≥18                     ║
║         DlPer≥35 | delt≤25 | Uptrend                                    ║
║  Base : wRSI(×1) ADX(×1) DlPer(×0.3) proximity(×0.5) pcon(×0.3)        ║
║         +15 Strong  +10 BBsqueeze  +8 PremZone  logAscr(×2)            ║
║  New  : +12 SUPERTd=+1  CMF×10  +8 SQZ_ON  +6 WILLR<-60               ║
║         +5 clos>EMA21  +4 Stoch setup                                   ║
╚══════════════════════════════════════════════════════════════════════════╝
    """)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    # ── resolve CSV path ──
    # 1. Ask for snapshot type
    print("Choose snapshot source:")
    print("1. snapshot.csv")
    print("2. snapshot_all.csv")
    choice = input("Enter choice [default 1]: ").strip()

    if choice == '2':
        suffix_flag = "_all"
        pattern = "*snapshot_all.csv"
    else:
        suffix_flag = ""
        pattern = "*snapshot.csv"

    # 2. Ask for data target
    print("\nOptions: press Enter to latest data, enter a specific date (DD-MM-YYYY),")
    print("a range (DD-MM-YYYY to DD-MM-YYYY), or a custom filename.")
    user_input = input("Target: ").strip()

    targets = []

    if not user_input:
        all_files = glob.glob(pattern)
        if choice != '2':
            all_files = [f for f in all_files if not f.endswith("_all.csv")]
        
        files = [f for f in all_files if re.match(r"^\d{2}-\d{2}-\d{2,4}", os.path.basename(f))]

        if not files:
            print(f"No files matching {pattern} found in current directory.")
            sys.exit(1)

        # Sort by the date embedded in the filename (DD-MM-YY)
        def get_file_date(fname):
            name = os.path.basename(fname)
            try:
                return datetime.strptime(name[:8], "%d-%m-%y")
            except ValueError:
                try:
                    return datetime.strptime(name[:10], "%d-%m-%Y")
                except ValueError:
                    return datetime.fromtimestamp(0)

        files.sort(key=get_file_date)
        targets.append(files[-1])

    elif "to" in user_input:
        # Handle range
        try:
            start_str, end_str = [d.strip() for d in user_input.split('to')]
            start_dt = datetime.strptime(start_str, '%d-%m-%Y')
            end_dt = datetime.strptime(end_str, '%d-%m-%Y')
            
            curr = start_dt
            while curr <= end_dt:
                fname = f"{curr.strftime('%d-%m-%y')}snapshot{suffix_flag}.csv"
                if os.path.exists(fname):
                    targets.append(fname)
                curr += timedelta(days=1)
        except Exception as e:
            print(f"Error parsing date range: {e}")
            sys.exit(1)

    elif os.path.exists(user_input):
        targets.append(user_input)

    else:
        # Try as single date
        try:
            dt = datetime.strptime(user_input, '%d-%m-%Y')
            fname = f"{dt.strftime('%d-%m-%y')}snapshot{suffix_flag}.csv"
            if os.path.exists(fname):
                targets.append(fname)
            else:
                print(f"File {fname} not found.")
                sys.exit(1)
        except ValueError:
            print(f"Invalid input format or file not found: {user_input}")
            sys.exit(1)

    if not targets:
        print("No valid snapshot files identified to process.")
        sys.exit(1)

    print_logic_summary()

    for csv_path in targets:
        snap_date = parse_date_from_filename(csv_path)
        df = load_csv(csv_path)
        print(f"\n[INFO] Processing {csv_path} — snapshot date: {snap_date}")

        # ── run screens ──
        intraday_results, intraday_pool = screen_intraday(df, INTRADAY_CFG)
        swing_results,   swing_pool    = screen_swing(df, SWING_CFG)

        print_section(f"INTRADAY CANDIDATES  [{snap_date}]",  intraday_results, intraday_pool)
        print_section(f"WEEKLY SWING CANDIDATES  [{snap_date}]", swing_results, swing_pool)

        # ── dual-timeframe confluence ──
        both = set(intraday_results["symb"]) & set(swing_results["symb"])
        if both:
            print(f"\n{'─'*70}")
            print(f"  ⭐ DUAL-TIMEFRAME CONFLUENCE: {', '.join(sorted(both))}")
            print(f"  Appear in BOTH lists — highest conviction picks.")
            print(f"{'─'*70}")

        # ── export ──
        out_dir  = os.path.dirname(os.path.abspath(csv_path)) or "."
        suffix = "_all" if os.path.basename(csv_path).endswith("snapshot_all.csv") else ""
        base_name = f"picks_{snap_date.replace(' ', '_')}{suffix}"
        
        intraday_tagged = intraday_results.assign(screener_type="Intraday")
        swing_tagged    = swing_results.assign(screener_type="Swing")
        combined_df = pd.concat([intraday_tagged, swing_tagged], ignore_index=True)

        apply_professional_formatting(os.path.join(out_dir, f"{base_name}.xlsx"), combined_df)
        print(f"[INFO] Results saved → {base_name}.xlsx (Formatted)")


if __name__ == "__main__":
    
  main()